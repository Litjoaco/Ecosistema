from django.shortcuts import render, redirect, get_object_or_404
from usuario.models import Usuario, RUBRO_CHOICES # Asegurarse de que Usuario está importado
from usuario.forms import AdminUsuarioForm, AyudanteUsuarioForm
from .models import Reunion, Encuesta, RespuestaEncuesta, SoporteTicket, TicketRespuesta
from .forms import ReunionForm, EncuestaForm, SoporteTicketAdminForm, TicketRespuestaForm
from django.urls import reverse
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, HttpRequest
from django.db.models import Q, F, Count, Avg, Sum
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import json

from django.contrib.auth.hashers import check_password
# Create your views here.
def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        usuario_id = request.session.get('usuario_id')
        if not usuario_id:
            return redirect('login')

        try:
            usuario_actual = Usuario.objects.get(id=usuario_id)
            # Ahora, ni admin ni ayudante pueden acceder a vistas solo para superadmin (si las hubiera)
            if not usuario_actual.es_admin and not usuario_actual.es_ayudante:
                return redirect('inicio')
        except Usuario.DoesNotExist:
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper
    
def solo_admin_required(view_func):
    """
    Decorador que restringe el acceso SOLO a los usuarios que son administradores (`es_admin`).
    """
    def wrapper(request, *args, **kwargs):
        usuario_id = request.session.get('usuario_id')
        if not usuario_id:
            return redirect('login')
        usuario_actual = get_object_or_404(Usuario, id=usuario_id)
        if not usuario_actual.es_admin:
            messages.error(request, "No tienes permiso para realizar esta acción.")
            return redirect('panel-admin:panel_admin') # Redirige al panel principal si no es admin
        return view_func(request, *args, **kwargs)
    return wrapper

def totem_required(view_func):
    """
    Decorador que restringe el acceso SOLO a los usuarios que son de tipo Tótem (`es_totem`).
    """
    def wrapper(request, *args, **kwargs):
        usuario_id = request.session.get('usuario_id')
        if not usuario_id:
            return redirect('login')
        usuario_actual = get_object_or_404(Usuario, id=usuario_id)
        if not usuario_actual.es_totem:
            messages.error(request, "Esta sección es solo para terminales de tipo Tótem.")
            return redirect('inicio')
        return view_func(request, *args, **kwargs)
    return wrapper

def privileged_user_required(view_func):
    """
    Decorador que permite el acceso a Administradores, Ayudantes y Tótems.
    Ideal para endpoints de API usados por diferentes roles.
    """
    def wrapper(request, *args, **kwargs):
        usuario_id = request.session.get('usuario_id')
        if not usuario_id:
            return redirect('login')
        try:
            usuario_actual = Usuario.objects.get(id=usuario_id)
            if not (usuario_actual.es_admin or usuario_actual.es_ayudante or usuario_actual.es_totem):
                return JsonResponse({'status': 'error', 'message': 'Permiso denegado.'}, status=403)
        except Usuario.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Usuario no encontrado.'}, status=403)
        return view_func(request, *args, **kwargs)
    return wrapper

@admin_required
def gestion_usuarios(request):
    usuario_actual = get_object_or_404(Usuario, id=request.session.get('usuario_id'))
    # La búsqueda y el filtrado ahora se manejan exclusivamente por AJAX.
    # Esta vista solo carga la página inicial con todos los usuarios.
    usuarios = Usuario.objects.all().order_by('nombre')
    rubros = Usuario.objects.exclude(rubro__exact='').values_list('rubro', flat=True).distinct().order_by('rubro')
    
    query = request.GET.get('q', '')
    rubro_filter = request.GET.get('rubro', '')

    usuarios = Usuario.objects.all().order_by('nombre', 'apellido')

    if query:
        usuarios = usuarios.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(email__icontains=query) |
            Q(rut__icontains=query)
        )
    
    if rubro_filter:
        usuarios = usuarios.filter(rubro=rubro_filter)

    return render(request, 'panel_admin_usuarios.html', {
        'usuarios': usuarios, 
        'rubros': rubros, 
        'query': query, 
        'rubro_filter': rubro_filter, 
        'usuario_actual': usuario_actual
    })

@admin_required
def buscar_usuarios_ajax(request):
    """
    Vista para manejar las peticiones AJAX de búsqueda de usuarios.
    """
    query = request.GET.get('q', '')
    rubro_filter = request.GET.get('rubro', '')

    usuarios = Usuario.objects.all().order_by('nombre', 'apellido')

    if query:
        usuarios = usuarios.filter(Q(nombre__icontains=query) | Q(apellido__icontains=query) | Q(email__icontains=query) | Q(rut__icontains=query))
    if rubro_filter:
        usuarios = usuarios.filter(rubro=rubro_filter)

    # Preparamos los datos para la respuesta JSON
    data = [{
        'id': u.id,
        'nombre': u.nombre,
        'apellido': u.apellido,
        'email': u.email,
        'rut': u.rut,
        'rubro': u.get_rubro_real_display,
        'telefono': u.telefono or '',
        'cantidad_asistencias': u.cantidad_asistencias,
        'es_admin': u.es_admin,
        'foto_url': u.foto.url if u.foto else '/static/img/person.jpg'
    } for u in usuarios]

    return JsonResponse({'usuarios': data})

@admin_required
def editar_usuario_admin(request, usuario_id):
    usuario_actual = get_object_or_404(Usuario, id=request.session.get('usuario_id'))
    usuario_a_editar = get_object_or_404(Usuario, id=usuario_id)

    # Determinar qué formulario usar según el rol del usuario actual
    if usuario_actual.es_admin:
        Formulario = AdminUsuarioForm
    else: # Si es ayudante
        Formulario = AyudanteUsuarioForm

    if request.method == 'POST':
        form = Formulario(request.POST, request.FILES, instance=usuario_a_editar)
        if form.is_valid():
            form.save()
            messages.success(request, f'¡Perfil de {usuario_a_editar.nombre} actualizado con éxito!')
            return redirect('panel-admin:gestion_usuarios')
    else:
        form = Formulario(instance=usuario_a_editar)

    return render(request, 'editar_perfil.html', {'form': form, 'usuario': usuario_a_editar})

@solo_admin_required
def eliminar_usuario(request, usuario_id):
    if request.method == 'POST':
        usuario_a_eliminar = get_object_or_404(Usuario, id=usuario_id)
        
        if usuario_a_eliminar.id == request.session.get('usuario_id'):
            messages.error(request, 'No puedes eliminar tu propia cuenta de administrador.')
            return redirect('panel-admin:gestion_usuarios')

        usuario_a_eliminar.delete()
        messages.success(request, f'Usuario {usuario_a_eliminar.email} eliminado correctamente.')
    # Redirigir al directorio si la acción vino de allí, si no, a la gestión de usuarios
    if 'from_directorio' in request.POST:
        return redirect('directorio_miembros')
    else:
        return redirect('gestion_usuarios')

@solo_admin_required
def toggle_destacado_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    usuario.destacado = not usuario.destacado
    usuario.save()
    return redirect('directorio_miembros')

@solo_admin_required
def toggle_visibilidad_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    # Un admin solo puede ocultar un perfil, no hacerlo público.
    if usuario.perfil_publico:
        usuario.perfil_publico = False
        usuario.save()
        messages.success(request, f'El perfil de {usuario.nombre} ha sido ocultado del directorio.')
    # Si ya es privado, no se hace nada.
    return redirect('directorio_miembros')

@solo_admin_required
def gestion_reuniones(request):
    if request.method == 'POST':
        form = ReunionForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Reunión creada con éxito!')
            return redirect('panel-admin:gestion_reuniones')
    else:
        form = ReunionForm()

    reuniones = Reunion.objects.all().order_by('-fecha')
    return render(request, 'panel_admin_reuniones.html', {
        'form': form,
        'reuniones': reuniones
    })

@solo_admin_required
def editar_reunion(request, reunion_id):
    reunion = get_object_or_404(Reunion, id=reunion_id)
    if request.method == 'POST':
        form = ReunionForm(request.POST, request.FILES, instance=reunion)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Reunión actualizada con éxito!')
            return redirect('panel-admin:gestion_reuniones')
    else:
        form = ReunionForm(instance=reunion)
    
    return render(request, 'panel_admin_reunion_editar.html', {'form': form, 'reunion': reunion})

@solo_admin_required
def eliminar_reunion(request, reunion_id):
    if request.method == 'POST':
        reunion = get_object_or_404(Reunion, id=reunion_id)
        reunion.delete()
        messages.success(request, 'Reunión eliminada correctamente.')
    return redirect('panel-admin:gestion_reuniones')

@admin_required # Ayudante puede ver la lista de reuniones para tomar asistencia
def control_asistencia(request):
    """
    Página que muestra un resumen de asistencia para todas las reuniones.
    """
    reuniones = Reunion.objects.all().order_by('-fecha')
    return render(request, 'panel_admin_control_asistencia.html', {
        'reuniones': reuniones
    })

@admin_required # Ayudante puede registrar asistencia
def registrar_asistencia(request, reunion_id):
    """
    Página para que el admin escanee QR y registre asistencia.
    """
    reunion = get_object_or_404(Reunion, id=reunion_id)
    
    if request.method == 'POST' and 'manual_add' in request.POST:
        usuario_id = request.POST.get('usuario_id')
        if usuario_id:
            usuario_a_agregar = get_object_or_404(Usuario, id=usuario_id)
            
            if not reunion.asistentes.filter(id=usuario_a_agregar.id).exists():
                reunion.asistentes.add(usuario_a_agregar)
                usuario_a_agregar.cantidad_asistencias = F('cantidad_asistencias') + 1
                usuario_a_agregar.save()
                
                # Si se debe imprimir etiqueta, preparamos la URL para la redirección.
                if reunion.imprimir_etiqueta_al_asistir:
                    redirect_url = f"{reverse('panel-admin:registrar_asistencia', args=[reunion_id])}?print_user={usuario_a_agregar.id}"
                    return redirect(redirect_url)
                else:
                    messages.success(request, f'Asistencia de {usuario_a_agregar.nombre} registrada manualmente.')
            else:
                messages.warning(request, f'{usuario_a_agregar.nombre} ya estaba registrado como asistente.')

            return redirect('panel-admin:registrar_asistencia', reunion_id=reunion_id)

    asistentes = reunion.asistentes.all().order_by('nombre')
    usuarios_no_asistentes = Usuario.objects.exclude(id__in=asistentes.values_list('id', flat=True)).order_by('nombre')

    return render(request, 'panel_admin_asistencia.html', {
        'reunion': reunion,
        'asistentes': asistentes,
        'usuarios_no_asistentes': usuarios_no_asistentes
    })

@admin_required # Ayudante puede quitar asistencia
def quitar_asistencia(request, reunion_id, usuario_id):
    """
    Elimina a un usuario de la lista de asistentes de una reunión.
    """
    if request.method == 'POST':
        reunion = get_object_or_404(Reunion, id=reunion_id)
        usuario_a_quitar = get_object_or_404(Usuario, id=usuario_id)

        if reunion.asistentes.filter(id=usuario_a_quitar.id).exists():
            reunion.asistentes.remove(usuario_a_quitar)
            # Decrementamos el contador de asistencias del usuario de forma segura
            if usuario_a_quitar.cantidad_asistencias > 0:
                usuario_a_quitar.cantidad_asistencias = F('cantidad_asistencias') - 1
                usuario_a_quitar.save()
            messages.success(request, f'Se ha quitado la asistencia de {usuario_a_quitar.nombre}.')
    
    return redirect('panel-admin:registrar_asistencia', reunion_id=reunion_id)

@privileged_user_required # Admin, Ayudante y Tótem pueden usar el QR para marcar asistencia
def marcar_asistencia_qr(request, reunion_id, usuario_id):
    """
    Endpoint API para marcar la asistencia de un usuario a una reunión.
    """
    if request.method == 'POST':
        reunion = get_object_or_404(Reunion, id=reunion_id)
        usuario = get_object_or_404(Usuario, id=usuario_id)

        # Primero, verificamos si el usuario ya es un asistente.
        if reunion.asistentes.filter(id=usuario.id).exists():
            # Si ya existe, devolvemos un error y no hacemos nada más.
            return JsonResponse({'status': 'error', 'message': f'{usuario.nombre} ya se encuentra registrado en esta reunión.'}, status=409)

        # Si no existe, procedemos a registrarlo.
        if not reunion.asistentes.filter(id=usuario.id).exists():
            reunion.asistentes.add(usuario)
            usuario.cantidad_asistencias = F('cantidad_asistencias') + 1
            usuario.save()

            # Refrescar el objeto para obtener el valor actualizado de 'cantidad_asistencias'
            usuario.refresh_from_db()

            print_url = reverse('imprimir_etiqueta', args=[usuario.id]) if reunion.imprimir_etiqueta_al_asistir else None
            return JsonResponse({
                'status': 'ok', 
                'message': f'Asistencia de {usuario.nombre} registrada.',
                'asistente': { 'id': usuario.id, 'nombre': usuario.nombre, 'apellido': usuario.apellido, 'rubro': usuario.get_rubro_real_display or '', 'foto_url': usuario.foto.url if usuario.foto else '/static/img/person.jpg' },
                'print_url': print_url
            })
            
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@solo_admin_required
def gestion_interesados(request):
    reuniones_proximas = Reunion.objects.filter(fecha__gte=timezone.now()).prefetch_related('interesados').order_by('fecha')
    return render(request, 'panel_admin_interesados.html', {
        'reuniones': reuniones_proximas
    })

@solo_admin_required
def gestion_encuestas(request):
    if request.method == 'POST':
        form = EncuestaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Encuesta creada con éxito.')
            return redirect('panel-admin:gestion_encuestas')
    else:
        form = EncuestaForm()

    encuestas = Encuesta.objects.all().order_by('-creada_en')
    return render(request, 'panel_admin_encuestas.html', {
        'form': form,
        'encuestas': encuestas
    })

@solo_admin_required
def ver_respuestas_encuesta(request, encuesta_id):
    encuesta = get_object_or_404(Encuesta, id=encuesta_id)
    respuestas = encuesta.respuestas.all().order_by('-fecha_respuesta')
    
    from django.db.models import Avg
    promedio = respuestas.aggregate(Avg('puntuacion'))['puntuacion__avg']

    return render(request, 'panel_admin_ver_respuestas.html', {
        'encuesta': encuesta,
        'respuestas': respuestas,
        'promedio': promedio
    })

@solo_admin_required
def eliminar_encuesta(request, encuesta_id):
    if request.method == 'POST':
        encuesta = get_object_or_404(Encuesta, id=encuesta_id)
        encuesta.delete()
        messages.success(request, 'Encuesta eliminada correctamente.')
    return redirect('panel-admin:gestion_encuestas')

@solo_admin_required
def toggle_destacado_respuesta(request, respuesta_id):
    if request.method == 'POST':
        respuesta = get_object_or_404(RespuestaEncuesta, id=respuesta_id)
        respuesta.destacado = not respuesta.destacado
        respuesta.save()
        messages.success(request, 'El estado de destacado del testimonio ha sido cambiado.')
        return redirect('panel-admin:ver_respuestas_encuesta', encuesta_id=respuesta.encuesta.id)
    return redirect('panel-admin:gestion_encuestas')

@admin_required # Ayudante puede ver y responder tickets
def gestion_soporte(request):
    tickets = SoporteTicket.objects.all().order_by('-fecha_creacion')
    return render(request, 'panel_admin_soporte.html', {'tickets': tickets})

@admin_required # Ayudante puede ver y responder tickets
def ver_ticket_soporte(request, ticket_id):
    ticket = get_object_or_404(SoporteTicket, id=ticket_id)
    admin_usuario = get_object_or_404(Usuario, id=request.session.get('usuario_id'))

    if request.method == 'POST':
        if 'actualizar_estado' in request.POST:
            estado_form = SoporteTicketAdminForm(request.POST, instance=ticket)
            if estado_form.is_valid():
                estado_form.save()
                messages.success(request, 'El estado del ticket ha sido actualizado.')
                return redirect('panel-admin:ver_ticket_soporte', ticket_id=ticket.id)
        
        if 'enviar_respuesta' in request.POST:
            respuesta_form = TicketRespuestaForm(request.POST, request.FILES)
            if respuesta_form.is_valid():
                respuesta = respuesta_form.save(commit=False)
                respuesta.ticket = ticket
                respuesta.usuario = admin_usuario
                respuesta.save()
                messages.success(request, 'Tu respuesta ha sido enviada.')
                return redirect('panel-admin:ver_ticket_soporte', ticket_id=ticket.id)

    estado_form = SoporteTicketAdminForm(instance=ticket)
    respuesta_form = TicketRespuestaForm()
    
    contexto = {
        'ticket': ticket,
        'estado_form': estado_form,
        'respuesta_form': respuesta_form,
        'admin_usuario': admin_usuario,
    }
    return render(request, 'panel_admin_ver_ticket.html', contexto)

@totem_required
def totem_seleccionar_reunion(request):
    """
    Vista para que el usuario Tótem seleccione la reunión que va a gestionar.
    """
    reuniones_proximas = Reunion.objects.filter(fecha__gte=timezone.now()).order_by('fecha')
    contexto = {
        'reuniones': reuniones_proximas,
        'usuario': get_object_or_404(Usuario, id=request.session.get('usuario_id'))
    }
    return render(request, 'totem_seleccionar_reunion.html', contexto)

@totem_required
def totem_escaner(request, reunion_id):
    """
    Vista de escaneo para el Tótem, bloqueada y a pantalla completa.
    """
    reunion = get_object_or_404(Reunion, id=reunion_id)
    # Pasamos el token CSRF explícitamente para que esté disponible en el JavaScript
    return render(request, 'totem_escaner.html', {'reunion': reunion})

@totem_required
def totem_verify_exit(request):
    """
    Endpoint API para verificar la contraseña del usuario Tótem y permitir la salida.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            password = data.get('password')
            usuario_totem = get_object_or_404(Usuario, id=request.session.get('usuario_id'))

            # Usamos check_password para comparar la contraseña en texto plano con la hasheada en la BD
            if check_password(password, usuario_totem.password): 
                return JsonResponse({'status': 'ok'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Contraseña incorrecta'}, status=401)
        except (json.JSONDecodeError, KeyError):
            return JsonResponse({'status': 'error', 'message': 'Petición inválida'}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@admin_required # Ayudante puede ver estadísticas (con lógica interna para restringir)
def estadisticas_admin(request):
    """
    Muestra la página de estadísticas, permitiendo una vista general
    o una vista detallada por reunión.
    """
    reuniones_para_filtro = Reunion.objects.all().order_by('-fecha')
    reunion_seleccionada_id_str = request.GET.get('reunion_id', None)
    usuario_actual = get_object_or_404(Usuario, id=request.session.get('usuario_id'))

    # --- 1. Función para aplanar los RUBRO_CHOICES anidados ---
    def flatten_choices(choices):
        flat_dict = {}
        for group in choices:
            if isinstance(group[1], (list, tuple)):
                for key, value in group[1]:
                    flat_dict[key] = value
            else:
                # En caso de que haya una tupla no anidada
                flat_dict[group[0]] = group[1]
        return flat_dict

    rubros_dict = flatten_choices(RUBRO_CHOICES)

    # --- 2. Inicialización completa del contexto con valores por defecto ---
    contexto = {
        'reuniones_para_filtro': reuniones_para_filtro,
        'usuario_actual': usuario_actual,
        'reunion_seleccionada_id': None,
        'reunion_seleccionada': None,
        'total_usuarios': 0,
        'total_reuniones': 0,
        'total_asistencias': 0,
        'promedio_satisfaccion': 0,
        'labels_reuniones': [],
        'data_asistencia': [],
        'data_conversion': [],
        'labels_puntuacion': [],
        'data_puntuacion': [],
        'labels_rubro': [],
        'data_rubro': [],
    }

    # --- 3. Determinar qué vista mostrar (general o por reunión) ---
    reunion_seleccionada_id = None
    # Si es ayudante y no se especifica reunión, se fuerza la primera de la lista.
    if usuario_actual.es_ayudante and not reunion_seleccionada_id_str and reuniones_para_filtro:
        reunion_seleccionada_id = reuniones_para_filtro.first().id
    elif reunion_seleccionada_id_str and reunion_seleccionada_id_str.isdigit():
        reunion_seleccionada_id = int(reunion_seleccionada_id_str)

    contexto['reunion_seleccionada_id'] = reunion_seleccionada_id

    # --- 4. Poblar el contexto con los datos correspondientes ---
    if reunion_seleccionada_id:
        # --- VISTA DE ESTADÍSTICAS POR REUNIÓN ---
        reunion = get_object_or_404(Reunion, id=reunion_seleccionada_id)
        contexto['reunion_seleccionada'] = reunion
        asistentes = reunion.asistentes.all()
        interesados = reunion.interesados.all()
        contexto['total_asistencias'] = asistentes.count()
        contexto['data_conversion'] = [interesados.count(), asistentes.count()]

        if hasattr(reunion, 'encuesta'):
            respuestas = reunion.encuesta.respuestas.all()
            contexto['promedio_satisfaccion'] = respuestas.aggregate(avg=Avg('puntuacion'))['avg'] or 0
            puntuaciones = respuestas.values('puntuacion').annotate(cantidad=Count('id')).order_by('puntuacion')
            contexto['labels_puntuacion'] = [f"{p['puntuacion']} Estrellas" for p in puntuaciones]
            contexto['data_puntuacion'] = [p['cantidad'] for p in puntuaciones]

        top_rubros = asistentes.filter(rubro__isnull=False).exclude(rubro__exact='').values('rubro').annotate(cantidad=Count('id')).order_by('-cantidad')[:5]
        contexto['labels_rubro'] = [rubros_dict.get(r['rubro'], r['rubro']) for r in top_rubros]
        contexto['data_rubro'] = [r['cantidad'] for r in top_rubros]

    elif usuario_actual.es_admin:
        # --- VISTA DE ESTADÍSTICAS GENERALES ---
        contexto['total_usuarios'] = Usuario.objects.count()
        contexto['total_reuniones'] = Reunion.objects.count()
        contexto['total_asistencias'] = Usuario.objects.aggregate(total=Sum('cantidad_asistencias'))['total'] or 0
        contexto['promedio_satisfaccion'] = RespuestaEncuesta.objects.aggregate(avg=Avg('puntuacion'))['avg'] or 0
        
        # Gráfico de Asistencia a últimas reuniones
        reuniones_recientes = Reunion.objects.annotate(num_asistentes=Count('asistentes')).order_by('-fecha')[:10][::-1] # Invertido para orden cronológico
        contexto['labels_reuniones'] = [r.detalle for r in reuniones_recientes]
        contexto['data_asistencia'] = [r.num_asistentes for r in reuniones_recientes]
        
        # Gráfico de Rubros (general)
        top_rubros = Usuario.objects.filter(rubro__isnull=False).exclude(rubro__exact='').values('rubro').annotate(cantidad=Count('id')).order_by('-cantidad')[:5]
        contexto['labels_rubro'] = [rubros_dict.get(r['rubro'], r['rubro']) for r in top_rubros]
        contexto['data_rubro'] = [r['cantidad'] for r in top_rubros]

    return render(request, 'panel_admin_estadisticas.html', contexto)

@admin_required # Ayudante puede exportar (con lógica interna para restringir)
def exportar_estadisticas_excel(request):
    """
    Exporta las estadísticas clave a un archivo Excel (.xlsx).
    Si se proporciona un `reunion_id`, exporta las estadísticas de esa reunión.
    De lo contrario, exporta las estadísticas generales.
    """
    reunion_id = request.GET.get('reunion_id')
    usuario_actual = get_object_or_404(Usuario, id=request.session.get('usuario_id'))
    workbook = openpyxl.Workbook()
    bold_font = Font(bold=True, size=12)
    filename = "estadisticas_ecosistemala.xlsx"

    if usuario_actual.es_ayudante and not reunion_id:
        messages.error(request, "No tienes permiso para exportar estadísticas generales.")
        return redirect('panel-admin:estadisticas_admin')

    if reunion_id and reunion_id.isdigit():
        # --- EXPORTAR ESTADÍSTICAS DE UNA REUNIÓN ESPECÍFICA ---
        reunion = get_object_or_404(Reunion, id=reunion_id)
        filename = f"estadisticas_{reunion.detalle.replace(' ', '_').lower()}_{reunion.fecha.strftime('%Y%m%d')}.xlsx"

        asistentes = reunion.asistentes.all()
        interesados = reunion.interesados.all()

        # Hoja de Resumen de la Reunión
        sheet_resumen = workbook.active
        sheet_resumen.title = "Resumen Reunión"
        sheet_resumen['A1'] = f"Estadísticas de la Reunión: {reunion.detalle}"
        sheet_resumen['A1'].font = Font(bold=True, size=14)

        promedio_satisfaccion = 0
        if hasattr(reunion, 'encuesta'):
            promedio_satisfaccion = reunion.encuesta.respuestas.aggregate(avg=Avg('puntuacion'))['avg'] or 0

        resumen_data = [
            ("Interesados", interesados.count()),
            ("Asistentes", asistentes.count()),
            ("Satisfacción Promedio", f"{promedio_satisfaccion:.2f} / 5" if promedio_satisfaccion else "N/A")
        ]
        for i, (label, value) in enumerate(resumen_data, start=3):
            sheet_resumen[f'A{i}'] = label
            sheet_resumen[f'B{i}'] = value
            sheet_resumen[f'A{i}'].font = bold_font

        # Hoja de Lista de Asistentes
        sheet_asistentes = workbook.create_sheet(title="Lista de Asistentes")
        sheet_asistentes.append(['Nombre', 'Apellido', 'Email', 'Rubro'])
        for cell in sheet_asistentes[1]: cell.font = bold_font
        for asistente in asistentes.order_by('nombre'):
            sheet_asistentes.append([asistente.nombre, asistente.apellido, asistente.email, asistente.get_rubro_real_display])

    elif usuario_actual.es_admin:
        # --- EXPORTAR ESTADÍSTICAS GENERALES (comportamiento actual) ---
        total_usuarios = Usuario.objects.count()
        total_reuniones = Reunion.objects.count()
        total_asistencias = Usuario.objects.aggregate(total=Sum('cantidad_asistencias'))['total'] or 0
        promedio_satisfaccion = RespuestaEncuesta.objects.aggregate(avg=Avg('puntuacion'))['avg'] or 0

        reuniones_asistencia = Reunion.objects.annotate(num_asistentes=Count('asistentes')).order_by('-fecha')
        distribucion_puntuacion = RespuestaEncuesta.objects.values('puntuacion').annotate(cantidad=Count('id')).order_by('puntuacion')
        usuarios_por_rubro_qs = Usuario.objects.filter(rubro__isnull=False).exclude(rubro__exact='').values('rubro').annotate(cantidad=Count('id')).order_by('-cantidad')
        
        def flatten_choices(choices):
            flat_dict = {}
            for group in choices:
                if isinstance(group[1], (list, tuple)):
                    for key, value in group[1]:
                        flat_dict[key] = value
                else:
                    flat_dict[group[0]] = group[1]
            return flat_dict
        rubros_dict = flatten_choices(RUBRO_CHOICES)

        # Hoja de Resumen General
        sheet_resumen = workbook.active
        sheet_resumen.title = "Resumen General"
        sheet_resumen['A1'] = "Estadísticas Generales"
        sheet_resumen['A1'].font = Font(bold=True, size=14)
        
        resumen_data = [
            ("Usuarios Totales", total_usuarios),
            ("Reuniones Totales", total_reuniones),
            ("Asistencias Totales", total_asistencias),
            ("Satisfacción Promedio", f"{promedio_satisfaccion:.2f} / 5" if promedio_satisfaccion else "N/A")
        ]
        for i, (label, value) in enumerate(resumen_data, start=3):
            sheet_resumen[f'A{i}'] = label
            sheet_resumen[f'B{i}'] = value
            sheet_resumen[f'A{i}'].font = bold_font

        # Hoja de Asistencia por Reunión
        sheet_asistencia = workbook.create_sheet(title="Asistencia por Reunión")
        sheet_asistencia.append(['Reunión', 'Fecha', 'Nº de Asistentes'])
        for cell in sheet_asistencia[1]: cell.font = bold_font
        for reunion in reuniones_asistencia:
            sheet_asistencia.append([reunion.detalle, reunion.fecha.strftime('%d-%m-%Y'), reunion.num_asistentes])

        # Hoja de Rubros
        sheet_rubros = workbook.create_sheet(title="Usuarios por Rubro")
        sheet_rubros.append(['Rubro', 'Cantidad de Usuarios'])
        for cell in sheet_rubros[1]: cell.font = bold_font
        for item in usuarios_por_rubro_qs:
            nombre_rubro = rubros_dict.get(item['rubro'], item['rubro'])
            sheet_rubros.append([nombre_rubro, item['cantidad']])

        # Hoja de Satisfacción
        sheet_satisfaccion = workbook.create_sheet(title="Distribución de Satisfacción")
        sheet_satisfaccion.append(['Puntuación (Estrellas)', 'Cantidad de Votos'])
        for cell in sheet_satisfaccion[1]: cell.font = bold_font
        for item in distribucion_puntuacion:
            sheet_satisfaccion.append([f"{item['puntuacion']} Estrellas", item['cantidad']])

    # Ajustar ancho de columnas en todas las hojas
    for sheet in workbook.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    # 3. Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)

    return response
